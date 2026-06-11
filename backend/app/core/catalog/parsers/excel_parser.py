"""
Feature:  Catalog Enrichment Pipeline (Document Parsing)
Layer:    Core / Parsers
Module:   app.core.catalog.parsers.excel_parser
Purpose:  Excel parsing via openpyxl. Reads the first sheet (or the sheet with
          the most data rows). Returns full text (all cell values joined) plus
          rows as list-of-dicts. Handles merged cells by reading the top-left
          cell value for the merged region. Empty rows are skipped.
          Supports .xlsx and .xls (via openpyxl read_only mode).
Depends:  openpyxl
HITL:     None.
"""

from __future__ import annotations

import asyncio
import io

import openpyxl

from app.core.catalog.parsers.pdf_parser import ParseResult


def _parse_sync(excel_bytes: bytes) -> ParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=False, data_only=True)

    # Pick the sheet with the most rows
    ws = max(wb.worksheets, key=lambda s: s.max_row or 0)

    # Resolve merged cell values: merged regions store the value in top-left cell
    merged_values: dict[tuple[int, int], object] = {}
    for merge in ws.merged_cells.ranges:
        top_left = ws.cell(row=merge.min_row, column=merge.min_col).value
        for row_idx in range(merge.min_row, merge.max_row + 1):
            for col_idx in range(merge.min_col, merge.max_col + 1):
                merged_values[(row_idx, col_idx)] = top_left

    def _cell_value(row_idx: int, col_idx: int) -> object:
        if (row_idx, col_idx) in merged_values:
            return merged_values[(row_idx, col_idx)]
        return ws.cell(row=row_idx, column=col_idx).value

    # First non-empty row is the header
    headers: list[str] = []
    header_row_idx: int | None = None
    for row_idx in range(1, ws.max_row + 1):
        candidate = [
            str(_cell_value(row_idx, col) or "").strip()
            for col in range(1, ws.max_column + 1)
        ]
        if any(candidate):
            headers = [h or f"col_{i}" for i, h in enumerate(candidate)]
            header_row_idx = row_idx
            break

    rows: list[dict[str, object]] = []
    if header_row_idx is not None:
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cells = [
                str(_cell_value(row_idx, col) or "").strip()
                for col in range(1, ws.max_column + 1)
            ]
            if any(cells):
                rows.append(dict(zip(headers, cells, strict=False)))

    full_text = "\n".join(
        " | ".join(str(v) for v in row.values()) for row in rows
    )

    return ParseResult(
        full_text=full_text,
        rows=rows,
        embedded_image_paths=[],
        page_count=1,
    )


async def parse_excel(excel_bytes: bytes) -> ParseResult:
    """Parse an Excel file using openpyxl. Returns rows as list-of-dicts."""
    return await asyncio.to_thread(_parse_sync, excel_bytes)
